#! python3
# writeExcel.py - 写入excel文档
import openpyxl
wb = openpyxl.load_workbook('.\\example.xlsx')
sheet = wb.active
sheet.title = 'Spam Spam Spam'
wb.save('.\\example_copy.xlsx')
print(wb.get_sheet_names)
wb.create_sheet()
wb.create_sheet(index=0,title='First Sheet')
print(wb.get_sheet_names)
wb.create_sheet(index=2,title='Middle Sheet')
print(wb.get_sheet_names)
wb.save('.\\example_copy2.xlsx')