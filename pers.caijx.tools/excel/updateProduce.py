#! python3
# updateProduce.py - 更新产品销售电子表格中的单元格
import openpyxl
wb = openpyxl.load_workbook('.\\produceSales.xlsx')
sheet = wb['Sheet']
# The product types and their updated prices
PRICE_UPDATES = {
    'Garlic':3.07,
    'Celery':1.19,
    'Lemon':1.27
}
# Loop through the rows and update the prices.
for rowNum in range(2,sheet.max_row): # skip the first row
    produceName = sheet.cell(row=rowNum,column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum,column=2).value = PRICE_UPDATES[produceName]
wb.save('.\\updatedProdeceSales.xlsx')