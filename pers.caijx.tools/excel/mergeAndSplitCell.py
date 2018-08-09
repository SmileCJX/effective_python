#! python3
# mergeAndSplitCell.py - 合并和拆分单元格
import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active
sheet.merge_cells('A1:D3')
sheet['A1'] = 'Twelve cells merged together.'
sheet.merge_cells('C5:D5')
sheet['C5'] = 'Two merged cells'
wb.save('.\\merged.xlsx')