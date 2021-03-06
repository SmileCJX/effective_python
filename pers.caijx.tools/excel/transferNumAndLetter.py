#! python3
# transferNumAndLetter.py - 列字母和数字之间的转换
import openpyxl
from openpyxl.utils import get_column_letter,column_index_from_string
print(get_column_letter(1))
print(get_column_letter(2))
print(get_column_letter(27))
print(get_column_letter(900))
wb = openpyxl.load_workbook('.\\example.xlsx')
sheet = wb['Sheet1']
print(get_column_letter(sheet.max_column))
print(column_index_from_string('A'))
print(column_index_from_string('AA'))