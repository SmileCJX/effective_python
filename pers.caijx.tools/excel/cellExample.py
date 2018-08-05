#! python3
# cellExample.py - 从表中取得单元格
import openpyxl
wb = openpyxl.load_workbook('.\\example.xlsx')
sheet = wb['Sheet1']
print(sheet['A1'])
print(sheet['A1'].value)
c = sheet['B1']
print(c.value)
print('Row ' + str(c.row) + ',Column ' + c.column + ' is ' + c.value)
print('Cell ' + c.coordinate + ' is ' + c.value)
print(sheet['C1'].value)

print(sheet.cell(row=1,column=2))
print(sheet.cell(row=1,column=2).value)
for i in range(1,8,2):
    print(i,sheet.cell(row=i,column=2).value)

# 确认表的大小
print(str(sheet.max_row))
print(str(sheet.max_column))